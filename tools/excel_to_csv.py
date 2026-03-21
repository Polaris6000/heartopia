"""
excel_to_csv.py
===============
Excel(.xlsx) 파일의 모든 시트를 각각 CSV로 변환하여 tools/csv/ 폴더에 저장하는 유틸리티.
csv_to_json.py 실행 전 선행 단계로 사용한다.

[전체 파이프라인]
  Excel(.xlsx)  →  excel_to_csv.py  →  tools/csv/*.csv  →  csv_to_json.py  →  src/data/*.json

[기본 동작]
  시트명 앞의 이모지를 제거한 뒤 CSV 파일명으로 사용.
  예: "🐦 새" → 새.csv / "🍳 레시피" → 레시피.csv

[사용법] (프로젝트 루트에서 실행)
  # 전체 시트 일괄 변환 (기본)
  python tools/excel_to_csv.py tools/excel/heartopia_data.xlsx

  # 특정 시트만 변환
  python tools/excel_to_csv.py tools/excel/heartopia_data.xlsx --sheet "🐦 새"

  # 특정 시트 + 출력 파일명 지정
  python tools/excel_to_csv.py tools/excel/heartopia_data.xlsx --sheet "🍳 레시피" --name 레시피.csv

  # 시트 목록 확인 (변환 없이 시트명만 출력)
  python tools/excel_to_csv.py tools/excel/heartopia_data.xlsx --list-sheets

참고:
  - xlsx 파일은 프로젝트 루트 기준 경로로 지정
  - 출력은 항상 tools/csv/ 폴더에 저장 (자동 덮어쓰기)
  - 의존성: pip install openpyxl
"""

import argparse
import csv
import re
from pathlib import Path

# openpyxl 임포트 — 없으면 안내 메시지 출력 후 종료
try:
    import openpyxl
except ImportError:
    print("❌ openpyxl이 설치되어 있지 않습니다.")
    print("   pip install openpyxl  으로 설치 후 다시 실행하세요.")
    raise SystemExit(1)

# ─────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────

# 이 스크립트 위치(tools/)의 상위 = 프로젝트 루트
TOOLS_DIR    = Path(__file__).resolve().parent   # tools/
PROJECT_ROOT = TOOLS_DIR.parent                  # heartopia/
CSV_DIR      = TOOLS_DIR / "csv"                 # tools/csv/


# ─────────────────────────────────────────
# 유틸리티
# ─────────────────────────────────────────

def clean_sheet_name(sheet_name: str) -> str:
    """
    시트명에서 이모지·특수문자를 제거하고 파일명으로 쓸 수 있는 문자열로 변환.

    변환 규칙:
      - 공백 기준으로 토큰 분리 후 한글/영문/숫자로만 구성된 토큰만 취합
      - 예: "🐦 새"     → "새"
      - 예: "🛍️ 상점"  → "상점"
      - 예: "Sheet1"   → "Sheet1"
    """
    tokens = sheet_name.strip().split()
    # 한글, 영문, 숫자, 밑줄만 포함된 토큰을 유효한 토큰으로 간주
    valid_tokens = [t for t in tokens if re.fullmatch(r'[\w가-힣]+', t)]
    return " ".join(valid_tokens) if valid_tokens else sheet_name.strip()


# ─────────────────────────────────────────
# 핵심 변환 함수
# ─────────────────────────────────────────

def print_sheet_list(xlsx_path: Path) -> None:
    """워크북의 시트 목록과 변환될 CSV 파일명을 출력한다."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"\n  [{xlsx_path.name}] 시트 목록:")
    for i, name in enumerate(wb.sheetnames, 1):
        csv_name = clean_sheet_name(name) + ".csv"
        print(f"    {i}. {name}  →  {csv_name}")
    wb.close()


def convert_sheet_to_csv(ws, out_path: Path) -> int:
    """
    워크시트 객체를 읽어 CSV 파일로 저장한다.

    Args:
        ws:       openpyxl 워크시트 객체
        out_path: 저장할 CSV 파일 경로

    Returns:
        저장된 데이터 행 수 (헤더 제외)
    """
    # 출력 디렉토리 생성 (없으면 자동 생성)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # CSV 쓰기 (한글 깨짐 방지: utf-8-sig = BOM 포함)
    row_count = 0
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # 모든 셀이 None인 완전 빈 행은 건너뜀
            if all(cell is None for cell in row):
                continue
            # None 셀은 빈 문자열로 변환
            writer.writerow(["" if cell is None else str(cell) for cell in row])
            if i > 0:  # 헤더(i=0) 제외한 데이터 행 카운트
                row_count += 1

    return row_count


def convert_all_sheets(xlsx_path: Path) -> None:
    """
    워크북의 모든 시트를 각각 CSV 파일로 변환한다.
    시트명에서 이모지를 제거한 이름을 CSV 파일명으로 사용.

    예: "🐦 새" → tools/csv/새.csv
    """
    # data_only=True: 수식 셀을 마지막 계산값으로 읽기
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"\n  [{xlsx_path.name}] — 전체 시트 변환")

    for sheet_name in wb.sheetnames:
        csv_name = clean_sheet_name(sheet_name) + ".csv"
        out_path = CSV_DIR / csv_name
        ws       = wb[sheet_name]

        row_count = convert_sheet_to_csv(ws, out_path)
        print(f"  ✅ {sheet_name:<20}→  {out_path.relative_to(PROJECT_ROOT)}  ({row_count}행)")

    wb.close()


def convert_single_sheet(xlsx_path: Path, sheet_name: str, out_name: str | None) -> None:
    """
    특정 시트 하나만 CSV로 변환한다.

    Args:
        xlsx_path:  xlsx 파일 경로
        sheet_name: 변환할 시트명
        out_name:   출력 CSV 파일명 (None이면 시트명 자동 변환)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 시트 존재 여부 확인
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        print(f"  ❌ 시트 '{sheet_name}'을 찾을 수 없습니다.")
        print(f"     사용 가능한 시트: {available}")
        wb.close()
        return

    # 출력 파일명 결정
    if out_name:
        csv_name = out_name if out_name.endswith(".csv") else f"{out_name}.csv"
    else:
        csv_name = clean_sheet_name(sheet_name) + ".csv"

    out_path  = CSV_DIR / csv_name
    ws        = wb[sheet_name]
    row_count = convert_sheet_to_csv(ws, out_path)
    print(f"\n  ✅ {sheet_name}  →  {out_path.relative_to(PROJECT_ROOT)}  ({row_count}행)")
    wb.close()


# ─────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Excel(.xlsx) 전체 시트 → CSV 변환기 (출력: tools/csv/)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "file",
        type=Path,
        help="변환할 xlsx 파일 경로 (프로젝트 루트 기준)",
    )
    parser.add_argument(
        "--sheet",
        metavar="SHEET",
        help="특정 시트만 변환 (미지정 시 전체 시트 변환)",
    )
    parser.add_argument(
        "--name",
        metavar="FILENAME",
        help="출력 CSV 파일명 지정 (예: 레시피.csv). --sheet 와 함께 사용.",
    )
    parser.add_argument(
        "--list-sheets",
        action="store_true",
        help="시트 목록과 변환될 CSV 파일명을 출력하고 종료",
    )

    args = parser.parse_args()

    # 상대 경로는 프로젝트 루트 기준으로 해석
    xlsx_path = (PROJECT_ROOT / args.file).resolve() if not args.file.is_absolute() else args.file.resolve()

    # 파일 존재 확인
    if not xlsx_path.exists():
        print(f"❌ 파일 없음: {xlsx_path}")
        raise SystemExit(1)

    print(f"\n{'─' * 50}")
    print(f"  heartopia Excel → CSV 변환기")
    print(f"  출력 폴더: {CSV_DIR.relative_to(PROJECT_ROOT)}/")
    print(f"{'─' * 50}")

    if args.list_sheets:
        # 시트 목록만 출력
        print_sheet_list(xlsx_path)

    elif args.sheet:
        # 특정 시트만 변환
        convert_single_sheet(xlsx_path, args.sheet, args.name)

    else:
        # 기본: 전체 시트 일괄 변환
        convert_all_sheets(xlsx_path)

    print(f"\n{'─' * 50}\n  완료!\n{'─' * 50}\n")


if __name__ == "__main__":
    main()
